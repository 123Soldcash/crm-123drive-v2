#!/usr/bin/env python3
"""
DealMachine Excel Converter
Converts DealMachine export Excel files to CRM-compatible format
"""

import openpyxl
import json
import sys
from pathlib import Path

# Column mapping from DealMachine names to CRM names
COLUMN_MAPPING = {
    # Basic info
    "property_id": "propertyId",
    "lead_id": "leadId",
    
    # Address
    "property_address_line_1": "addressLine1",
    "property_address_line_2": "addressLine2",
    "property_address_city": "city",
    "property_address_state": "state",
    "property_address_zipcode": "zipcode",
    "property_address_county": "county",
    "property_lat": "latitude",
    "property_lng": "longitude",
    
    # Owner 1
    "owner_1_name": "owner1Name",
    "owner_1_address": "owner1Address",
    "owner_1_city": "owner1City",
    "owner_1_state": "owner1State",
    "owner_1_zipcode": "owner1Zipcode",
    
    # Owner 2
    "owner_2_name": "owner2Name",
    "owner_2_address": "owner2Address",
    "owner_2_city": "owner2City",
    "owner_2_state": "owner2State",
    "owner_2_zipcode": "owner2Zipcode",
    
    # Mailing address
    "mailing_address_line_1": "mailingAddressLine1",
    "mailing_address_line_2": "mailingAddressLine2",
    "mailing_city": "mailingCity",
    "mailing_state": "mailingState",
    "mailing_zipcode": "mailingZipcode",
    
    # Property details
    "property_type": "propertyType",
    "beds": "totalBedrooms",
    "baths": "totalBaths",
    "sqft": "buildingSquareFeet",
    "lot_sqft": "lotSqft",
    "year_built": "yearBuilt",
    "property_condition": "propertyCondition",
    "roof_age": "roofAge",
    "roof_type": "roofType",
    "ac_age": "acAge",
    "ac_type": "acType",
    "foundation_type": "foundationType",
    "pool": "pool",
    "garage_type": "garageType",
    "garage_spaces": "garageSpaces",
    "stories": "stories",
    "hoa_fee": "hoaFee",
    
    # Financial
    "property_tax_amount": "taxAmount",
    "estimated_value": "estimatedValue",
    "equity_amount": "equityAmount",
    "equity_percent": "equityPercent",
    "days_on_market": "daysOnMarket",
    "last_sale_date": "lastSaleDate",
    "last_sale_price": "lastSalePrice",
    "occupancy_status": "occupancyStatus",
    "mls_status": "mlsStatus",
    "lease_type": "leaseType",
    
    # Mortgages
    "first_mortgage_lender": "mtg1Lender",
    "first_mortgage_amount": "mtg1LoanAmt",
    "mtg1_est_loan_balance": "mtg1EstLoanBalance",
    "mtg1_est_payment_amount": "mtg1EstPaymentAmount",
    "first_mortgage_loan_type": "mtg1LoanType",
    "first_mortgage_financing_type": "mtg1TypeFinancing",
    "first_mortgage_interest_rate": "mtg1EstInterestRate",
    
    "second_mortgage_lender": "mtg2Lender",
    "second_mortgage_amount": "mtg2LoanAmt",
    "mtg2_est_loan_balance": "mtg2EstLoanBalance",
    "mtg2_est_payment_amount": "mtg2EstPaymentAmount",
    "second_mortgage_loan_type": "mtg2LoanType",
    "second_mortgage_financing_type": "mtg2TypeFinancing",
    "second_mortgage_interest_rate": "mtg2EstInterestRate",
    
    "mtg3_lender": "mtg3Lender",
    "mtg3_loan_amt": "mtg3LoanAmt",
    "mtg3_est_loan_balance": "mtg3EstLoanBalance",
    "mtg3_est_payment_amount": "mtg3EstPaymentAmount",
    "mtg3_loan_type": "mtg3LoanType",
    "mtg3_type_financing": "mtg3TypeFinancing",
    "mtg3_est_interest_rate": "mtg3EstInterestRate",
    
    "mtg4_lender": "mtg4Lender",
    "mtg4_loan_amt": "mtg4LoanAmt",
    "mtg4_est_loan_balance": "mtg4EstLoanBalance",
    "mtg4_est_payment_amount": "mtg4EstPaymentAmount",
    "mtg4_loan_type": "mtg4LoanType",
    "mtg4_type_financing": "mtg4TypeFinancing",
    "mtg4_est_interest_rate": "mtg4EstInterestRate",
    
    # HOA
    "hoa_fee_amount": "hoaFeeAmount",
    "h_o_a1_name": "hoa1Name",
    "h_o_a1_type": "hoa1Type",
    "mail": "mail",
    "dealmachine_url": "dealmachineUrl",
    
    # Notes
    "notes_1": "notes1",
    "notes_2": "notes2",
    "notes_3": "notes3",
    "notes_4": "notes4",
    "notes_5": "notes5",
    
    # Social
    "facebookprofile1": "facebookProfile1",
    "facebookprofile2": "facebookProfile2",
    "facebookprofile3": "facebookProfile3",
    "facebookprofile4": "facebookProfile4",
    
    # Research
    "priority": "priority",
    "skiptracetruepeoplesearch": "skiptraceTrue",
    "calledtruepeoplesearch": "calledTrue",
    "done_with_facebook": "doneWithFacebook",
    "address_of_the_property": "addressOfProperty",
    "donemailing_-_onwers": "doneMailingOwners",
    "donemailingrelatives": "doneMailingRelatives",
    "emailonwersinstantly.ai": "emailOwnersInstantly",
    "idi_-_search": "idiSearch",
    "httpsofficialrecords.broward.orgacclaimwebsearchsearchtypename": "officialRecordsSearch",
    "httpscounty-taxes.netbrowardbrowardproperty-tax": "countyTaxSearch",
    "violationsearch": "violationSearch",
    "httpsofficialrecords.broward.orgacclaimwebsearchsearchtypesimplesearch": "simpleSearch",
    "httpsdpepp.broward.orgbcsdefault.aspxpossepresentationparcelpermitlistposseobjectid116746": "permitSearch",
    "skiptracemanus": "skiptraceManus",
    "calledmanus": "calledManus",
    "property_flags": "propertyFlags",
}

def convert_file(input_path, output_path):
    """Convert DealMachine Excel to CRM format"""
    
    print(f"📂 Opening: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f"✅ Found {len(headers)} columns")
    
    # Create new workbook
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Write new headers
    new_headers = []
    for h in headers:
        new_h = COLUMN_MAPPING.get(h, h)
        new_headers.append(new_h)
    
    new_ws.append(new_headers)
    print(f"✅ Mapped headers")
    
    # Copy data rows
    row_count = 0
    for row_idx in range(2, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col_idx)
            row_data.append(cell.value)
        new_ws.append(row_data)
        row_count += 1
    
    print(f"✅ Copied {row_count} data rows")
    
    # Save new file
    new_wb.save(output_path)
    print(f"✅ Saved to: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 convert-dealmachine-excel.py <input_file> [output_file]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace(".xlsx", "-converted.xlsx")
    
    convert_file(input_file, output_file)
    print(f"\n✨ Conversion complete!")
    print(f"Use the converted file: {output_file}")
