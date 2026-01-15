#!/usr/bin/env python3
"""
Analyze the CORRECT DealMachine Excel file structure
"""

import openpyxl
import json

# Load the correct Excel file
wb = openpyxl.load_workbook('/home/ubuntu/crm-123drive-v2/dealmachine-properties-2026-01-12-220953_rolando_test.xlsx')
sheet = wb.active

# Get all column names (first row)
columns = []
for cell in sheet[1]:
    if cell.value:
        columns.append(str(cell.value))

print(f"📊 CORRECT EXCEL FILE ANALYSIS")
print(f"=" * 80)
print(f"Total columns: {len(columns)}")
print(f"Total rows: {sheet.max_row - 1}")  # Exclude header
print()

# Categorize columns
lead_fields = []
property_fields = []
financial_fields = []
owner_fields = []
contact_fields = []
other_fields = []

for col in columns:
    col_lower = col.lower()
    
    if col_lower.startswith('contact_'):
        contact_fields.append(col)
    elif 'lead' in col_lower or col_lower == 'owner_1_name':
        lead_fields.append(col)
    elif any(x in col_lower for x in ['address', 'property', 'lat', 'lng', 'county', 'type', 'bedroom', 'bathroom', 'sqft', 'square', 'year', 'built']):
        property_fields.append(col)
    elif any(x in col_lower for x in ['value', 'equity', 'mortgage', 'loan', 'sale', 'price', 'tax', 'assessed']):
        financial_fields.append(col)
    elif 'owner' in col_lower:
        owner_fields.append(col)
    else:
        other_fields.append(col)

print("📋 FIELD CATEGORIES:")
print()
print(f"🏷️  LEAD FIELDS ({len(lead_fields)}):")
for field in lead_fields:
    print(f"  - {field}")
print()

print(f"🏠 PROPERTY FIELDS ({len(property_fields)}):")
for field in property_fields:
    print(f"  - {field}")
print()

print(f"💰 FINANCIAL FIELDS ({len(financial_fields)}):")
for field in financial_fields:
    print(f"  - {field}")
print()

print(f"👤 OWNER FIELDS ({len(owner_fields)}):")
for field in owner_fields:
    print(f"  - {field}")
print()

print(f"📞 CONTACT FIELDS ({len(contact_fields)}):")
# Group by contact number
contact_nums = set()
for field in contact_fields:
    parts = field.split('_')
    if len(parts) >= 2 and parts[1].isdigit():
        contact_nums.add(int(parts[1]))

print(f"  Total contact slots: {max(contact_nums) if contact_nums else 0}")
print(f"  Fields per contact: {len([f for f in contact_fields if f.startswith('contact_1_')])}")
print()
print(f"  Sample contact_1 fields:")
for field in sorted([f for f in contact_fields if f.startswith('contact_1_')]):
    print(f"    - {field}")
print()

print(f"🔧 OTHER FIELDS ({len(other_fields)}):")
for field in other_fields:
    print(f"  - {field}")
print()

# Sample data from first property
print("=" * 80)
print("📝 SAMPLE DATA (First Property):")
print("=" * 80)

if sheet.max_row > 1:
    row_data = {}
    for idx, cell in enumerate(sheet[2], 1):
        if idx <= len(columns):
            col_name = columns[idx-1]
            value = cell.value
            if value is not None and str(value).strip():
                row_data[col_name] = value
    
    print()
    print("LEAD INFO:")
    for field in lead_fields:
        if field in row_data:
            print(f"  {field}: {row_data[field]}")
    
    print()
    print("PROPERTY INFO:")
    for field in property_fields[:10]:  # First 10 property fields
        if field in row_data:
            print(f"  {field}: {row_data[field]}")
    
    print()
    print("FINANCIAL INFO:")
    for field in financial_fields[:10]:  # First 10 financial fields
        if field in row_data:
            print(f"  {field}: {row_data[field]}")
    
    print()
    print("CONTACT INFO (contact_1):")
    for field in sorted([f for f in contact_fields if f.startswith('contact_1_')]):
        if field in row_data:
            print(f"  {field}: {row_data[field]}")

print()
print("=" * 80)
print("✅ ANALYSIS COMPLETE")
print("=" * 80)
